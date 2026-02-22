"""
services/excel_import_service.py
═══════════════════════════════════════════════════════════════════════
سرویس ایمپورت اکسل بازیکنان
Excel Player Import Service

Reads the existing club Excel database (columns A–V, RTL, multiple sheets)
and upserts Player records into Django, resolving Jalali dates,
insurance cell colours, and auto-creating TrainingCategory objects.

Column layout (1-indexed, as seen in screenshot):
  A=1  ردیف          (row number — skip)
  B=2  نام           first_name
  C=3  نام خانوادگی  last_name
  D=4  نام پدر       father_name
  E=5  تاریخ تولد    dob  (Jalali "1389/09/01")
  F=6  شماره ملی     national_id  ← UNIQUE KEY
  G=7  همراه بازیکن  phone
  H=8  همراه پدر     father_phone
  I=9  همراه مادر    mother_phone
  J=10 اعتبار بیمه   insurance_expiry (date string) + cell colour
  K=11 سطح فنی       skill_level ("A","B","C"…)
  L=12 رده سنی       age_category label (display only)
  M=13 دسته تمرینی   TrainingCategory.name
  N=14 فرم ثبت نام   has_registration_form (دارد/ندارد)
  O=15 تحصیلات پدر   father_education
  P=16 تحصیلات مادر  mother_education
  Q=17 شغل پدر       father_job
  R=18 شغل مادر      mother_job
  S=19 قد             height (cm int)
  T=20 وزن            weight (kg decimal)
  U=21 دست            preferred_hand (راست=R / چپ=L)
  V=22 پا             preferred_foot (راست=R / چپ=L)
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import jdatetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

logger = logging.getLogger(__name__)


# ══════════════════════════════════════════════════════════════════════
#  CONSTANTS
# ══════════════════════════════════════════════════════════════════════

# Columns by 0-based index in the DataFrame (after dropping index col A)
COL = {
    "row_num":     0,   # A — skip
    "first_name":  1,   # B
    "last_name":   2,   # C
    "father_name": 3,   # D
    "dob":         4,   # E
    "national_id": 5,   # F
    "phone":       6,   # G
    "father_phone":7,   # H
    "mother_phone":8,   # I
    "insurance":   9,   # J
    "skill_level": 10,  # K
    "age_cat":     11,  # L
    "category":    12,  # M
    "reg_form":    13,  # N
    "father_edu":  14,  # O
    "mother_edu":  15,  # P
    "father_job":  16,  # Q
    "mother_job":  17,  # R
    "height":      18,  # S
    "weight":      19,  # T
    "hand":        20,  # U
    "foot":        21,  # V
}

# Education label → model choice key
EDUCATION_MAP = {
    "بی سواد": "illiterate", "بی‌سواد": "illiterate",
    "ابتدایی": "elementary",
    "راهنمایی": "middle",
    "سیکل":    "middle",
    "دیپلم":   "high_school",
    "فوق دیپلم": "associate", "فوق‌دیپلم": "associate",
    "لیسانس":  "bachelor",
    "کارشناسی": "bachelor",  "کارشناسی ارشد": "master",
    "فوق لیسانس": "master",  "فوق‌لیسانس": "master",
    "دکترا":   "phd",        "دکتری": "phd",
    "نظامی":   "other",
}

# Insurance cell fill colours (openpyxl hex, no '#')
INSURANCE_RED    = {"FFFF0000", "FFFA0000", "FFDC143C", "FFCC0000"}  # expired
INSURANCE_YELLOW = {"FFFFFF00", "FFFFD700", "FFFFC107", "FFFFF44F"}  # near expiry
INSURANCE_GREEN  = {"FF00FF00", "FF008000", "FF00B050", "FF92D050",
                    "FF70AD47", "FF00FF7F"}  # active

# Sheets to skip (non-player sheets)
SKIP_SHEETS = {"راهنما", "توضیحات", "فرمول", "Sheet1", "Sheet2", "Sheet3"}


# ══════════════════════════════════════════════════════════════════════
#  RESULT DATA CLASSES
# ══════════════════════════════════════════════════════════════════════

@dataclass
class RowResult:
    row_num:    int
    national_id: str
    name:       str
    action:     str           # "created" | "updated" | "skipped" | "error"
    message:    str = ""
    sheet:      str = ""


@dataclass
class ImportResult:
    total_rows:         int = 0
    created:            int = 0
    updated:            int = 0
    skipped:            int = 0
    errors:             int = 0
    categories_created: int = 0
    rows: List[RowResult] = field(default_factory=list)
    warnings: List[str]   = field(default_factory=list)

    @property
    def success_rate(self) -> float:
        if self.total_rows == 0:
            return 0.0
        return round((self.created + self.updated) / self.total_rows * 100, 1)


# ══════════════════════════════════════════════════════════════════════
#  JALALI DATE CONVERSION  (robust — handles multiple formats)
# ══════════════════════════════════════════════════════════════════════

_PERSIAN_TO_LATIN = str.maketrans("۰۱۲۳۴۵۶۷۸۹", "0123456789")


def _normalize_date_str(raw: str) -> str:
    """Convert Persian digits → Latin, strip whitespace, unify separators."""
    return raw.translate(_PERSIAN_TO_LATIN).strip().replace("-", "/").replace(".", "/")


def jalali_to_gregorian(raw) -> Optional[object]:
    """
    Convert a Jalali date value (string or pandas Timestamp) to a
    Python datetime.date or jdatetime.date that Django can store.

    Supported input formats:
        "1389/09/01"   — standard slash
        "1389-09-01"   — dash
        "13890901"     — compact
        "۱۳۸۹/۰۹/۰۱"  — Persian digits
        pandas NaT / NaN / None → returns None
    """
    import datetime

    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return None
    if isinstance(raw, (datetime.date, datetime.datetime)):
        return raw.date() if isinstance(raw, datetime.datetime) else raw

    raw_str = _normalize_date_str(str(raw))

    # Remove any non-numeric/slash residue
    raw_str = re.sub(r"[^\d/]", "", raw_str)

    if not raw_str or raw_str == "0" or len(raw_str) < 6:
        return None

    try:
        if "/" in raw_str:
            parts = raw_str.split("/")
        elif len(raw_str) == 8:
            parts = [raw_str[:4], raw_str[4:6], raw_str[6:]]
        else:
            return None

        if len(parts) != 3:
            return None

        y, m, d = int(parts[0]), int(parts[1]), int(parts[2])

        # Sanity check: Jalali year should be 1300–1420 for reasonable player ages
        if not (1300 <= y <= 1420):
            logger.warning("Unusual Jalali year: %d — skipping date", y)
            return None
        if not (1 <= m <= 12) or not (1 <= d <= 31):
            return None

        jdate = jdatetime.date(y, m, d)
        return jdate.togregorian()

    except (ValueError, TypeError, AttributeError) as e:
        logger.debug("Jalali parse failed for '%s': %s", raw, e)
        return None


def jalali_str_to_gregorian(raw) -> Optional[object]:
    """Alias kept for backwards compat."""
    return jalali_to_gregorian(raw)


# ══════════════════════════════════════════════════════════════════════
#  INSURANCE CELL DETECTOR
# ══════════════════════════════════════════════════════════════════════

def _normalise_hex(hex_str: Optional[str]) -> Optional[str]:
    if not hex_str:
        return None
    h = hex_str.upper().lstrip("#")
    return h if len(h) == 8 else ("FF" + h if len(h) == 6 else None)


@dataclass
class InsuranceInfo:
    status:      str                   # "active" | "expired" | "near_expiry" | "none"
    expiry_date: Optional[object] = None


def detect_insurance(cell_value, fill_color: Optional[str]) -> InsuranceInfo:
    """
    Determine insurance status from:
    1. Cell fill colour (red/yellow/green)
    2. Cell value (Jalali date string)

    Insurance column (J) rules observed in the screenshot:
    - Red fill      → insurance expired
    - Yellow fill   → insurance near expiry (date may be inside)
    - Has a date    → active insurance, date = expiry
    - Empty/no fill → no insurance
    """
    hex_norm = _normalise_hex(fill_color)

    # Determine colour-based status
    colour_status = "none"
    if hex_norm:
        if hex_norm in INSURANCE_RED:
            colour_status = "expired"
        elif hex_norm in INSURANCE_YELLOW:
            colour_status = "near_expiry"
        elif hex_norm in INSURANCE_GREEN:
            colour_status = "active"

    # Parse date value if present
    expiry_date = None
    if cell_value and str(cell_value).strip() not in ("", "nan", "None"):
        expiry_date = jalali_to_gregorian(cell_value)

    # Determine final status
    if expiry_date:
        import datetime
        today = datetime.date.today()
        if expiry_date < today:
            status = "expired"
        elif (expiry_date - today).days <= 30:
            status = "near_expiry"
        else:
            status = "active"
    elif colour_status != "none":
        status = colour_status
    else:
        status = "none"

    return InsuranceInfo(status=status, expiry_date=expiry_date)


# ══════════════════════════════════════════════════════════════════════
#  PHONE NORMALISER
# ══════════════════════════════════════════════════════════════════════

def normalise_phone(raw) -> str:
    """Normalise Iranian mobile numbers to 09xxxxxxxxx format."""
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return ""
    s = str(raw).translate(_PERSIAN_TO_LATIN).strip()
    # Remove any non-digit characters except leading +
    digits = re.sub(r"[^\d]", "", s)
    if len(digits) == 10 and digits.startswith("9"):
        return "0" + digits
    if len(digits) == 11 and digits.startswith("09"):
        return digits
    if len(digits) == 12 and digits.startswith("989"):
        return "0" + digits[2:]
    # Return cleaned as-is (validation happens in Django form)
    return digits[:11] if len(digits) >= 11 else digits


# ══════════════════════════════════════════════════════════════════════
#  NATIONAL ID NORMALISER
# ══════════════════════════════════════════════════════════════════════

def normalise_national_id(raw) -> Optional[str]:
    """Clean and validate Iranian national ID (10 digits)."""
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return None
    s = str(raw).translate(_PERSIAN_TO_LATIN).strip()
    digits = re.sub(r"[^\d]", "", s)

    # Handle scientific notation from Excel (e.g. "4.581E+9")
    if "E" in s.upper() or "e" in s:
        try:
            digits = str(int(float(s)))
        except ValueError:
            return None

    if len(digits) != 10:
        # Only pad 9-digit IDs (leading zero lost by Excel) — reject anything shorter
        if len(digits) == 9:
            digits = digits.zfill(10)
        else:
            return None

    return digits


# ══════════════════════════════════════════════════════════════════════
#  EDUCATION / HAND / FOOT MAPPERS
# ══════════════════════════════════════════════════════════════════════

def map_education(raw) -> str:
    if not raw or (isinstance(raw, float) and pd.isna(raw)):
        return ""
    s = str(raw).strip()
    return EDUCATION_MAP.get(s, "other" if s else "")


def map_hand_foot(raw) -> str:
    if not raw or (isinstance(raw, float) and pd.isna(raw)):
        return "R"
    s = str(raw).strip()
    return "L" if "چپ" in s else "R"


def safe_int(raw) -> Optional[int]:
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return None
    try:
        return int(float(str(raw).translate(_PERSIAN_TO_LATIN)))
    except (ValueError, TypeError):
        return None


def safe_decimal(raw) -> Optional[Decimal]:
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return None
    try:
        return Decimal(str(float(str(raw).translate(_PERSIAN_TO_LATIN))))
    except (InvalidOperation, ValueError, TypeError):
        return None


# ══════════════════════════════════════════════════════════════════════
#  CELL COLOUR EXTRACTOR
# ══════════════════════════════════════════════════════════════════════

def _extract_cell_fills(filepath: str, sheet_name: str, col_idx: int) -> Dict[int, Optional[str]]:
    """
    Read fill colours from a specific column using openpyxl (not pandas).
    Returns {row_number: hex_fill_color_or_None}
    col_idx is 1-based (Excel column number).
    """
    wb = load_workbook(filepath, read_only=False, data_only=True)
    if sheet_name not in wb.sheetnames:
        return {}

    ws = wb[sheet_name]
    colours: Dict[int, Optional[str]] = {}

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):  # skip header
        cell = row[col_idx - 1] if len(row) >= col_idx else None
        if cell is None:
            colours[row_idx] = None
            continue
        try:
            fill = cell.fill
            if fill and fill.fgColor and fill.fgColor.type == "rgb":
                colours[row_idx] = fill.fgColor.rgb
            elif fill and fill.fgColor and fill.fgColor.type == "theme":
                colours[row_idx] = None  # theme colour — treat as none
            else:
                colours[row_idx] = None
        except Exception:
            colours[row_idx] = None

    wb.close()
    return colours


# ══════════════════════════════════════════════════════════════════════
#  MAIN IMPORT SERVICE
# ══════════════════════════════════════════════════════════════════════

class ExcelImportService:
    """
    Reads an Excel workbook with one or more player sheets
    and upserts Django Player records.

    Usage:
        svc = ExcelImportService(filepath="/uploads/players.xlsx")
        result = svc.run(created_by=request.user, dry_run=False)
    """

    INSURANCE_COL_LETTER = "J"
    INSURANCE_COL_NUM    = 10   # 1-based

    def __init__(
        self,
        filepath: str,
        sheet_names: Optional[List[str]] = None,
        header_row: int = 0,
    ):
        self.filepath   = str(filepath)
        self.sheet_names = sheet_names   # None = all sheets
        self.header_row  = header_row    # 0-based for pandas

    # ── Public entry point ─────────────────────────────────────────
    def run(
        self,
        created_by=None,
        dry_run: bool = False,
    ) -> ImportResult:
        """
        Process the workbook and return an ImportResult summary.

        dry_run=True → parse and validate only, no DB writes.
        """
        result = ImportResult()

        # Discover which sheets to process
        xf = pd.ExcelFile(self.filepath)
        sheets_to_process = self.sheet_names or xf.sheet_names
        sheets_to_process = [s for s in sheets_to_process if s not in SKIP_SHEETS]

        logger.info("Excel import started: %s (%d sheets)", self.filepath, len(sheets_to_process))

        for sheet_name in sheets_to_process:
            try:
                self._process_sheet(sheet_name, result, created_by, dry_run)
            except Exception as exc:
                result.warnings.append(f"Sheet «{sheet_name}» skipped: {exc}")
                logger.exception("Failed processing sheet %s", sheet_name)

        logger.info(
            "Import complete: total=%d created=%d updated=%d errors=%d",
            result.total_rows, result.created, result.updated, result.errors
        )
        return result

    # ── Sheet processor ────────────────────────────────────────────
    def _process_sheet(
        self,
        sheet_name: str,
        result: ImportResult,
        created_by,
        dry_run: bool,
    ):
        # Load with pandas (fast)
        df = pd.read_excel(
            self.filepath,
            sheet_name=sheet_name,
            header=self.header_row,
            dtype=str,           # everything as string to avoid type coercion
        )

        # Skip obviously empty sheets
        if df.empty or len(df.columns) < 6:
            result.warnings.append(f"Sheet «{sheet_name}» skipped: not enough columns ({len(df.columns)})")
            return

        # Pre-extract insurance fill colours for this sheet (openpyxl pass)
        try:
            insurance_fills = _extract_cell_fills(
                self.filepath, sheet_name, self.INSURANCE_COL_NUM
            )
        except Exception as e:
            insurance_fills = {}
            result.warnings.append(f"Sheet «{sheet_name}»: could not read cell colours ({e})")

        # Process each row
        for df_idx, row in df.iterrows():
            result.total_rows += 1
            # openpyxl row number = df_idx + 2 (1 for header + 1 for 1-based)
            opx_row = int(df_idx) + 2

            rr = self._process_row(
                row=row,
                row_num=opx_row,
                sheet_name=sheet_name,
                insurance_fill=insurance_fills.get(opx_row),
                created_by=created_by,
                dry_run=dry_run,
                result=result,
            )
            result.rows.append(rr)

            if rr.action == "created":  result.created  += 1
            elif rr.action == "updated": result.updated  += 1
            elif rr.action == "skipped": result.skipped  += 1
            elif rr.action == "error":   result.errors   += 1

    # ── Row processor ──────────────────────────────────────────────
    def _process_row(
        self,
        row: pd.Series,
        row_num: int,
        sheet_name: str,
        insurance_fill: Optional[str],
        created_by,
        dry_run: bool,
        result: ImportResult,
    ) -> RowResult:

        def cell(idx: int):
            """Safe cell value getter by 0-based column index."""
            try:
                v = row.iloc[idx]
                return None if pd.isna(v) else str(v).strip()
            except (IndexError, TypeError):
                return None

        # ── 1. National ID (dedup key) ─────────────────────────────
        national_id = normalise_national_id(cell(COL["national_id"]))
        _nid_auto_generated = False
        if not national_id:
            # کد ملی خالی یا ناقص → شناسه موقت بر اساس نام + ردیف
            fn  = (cell(COL["first_name"]) or "").strip()
            ln  = (cell(COL["last_name"])  or "").strip()
            if not fn and not ln:
                return RowResult(
                    row_num=row_num, national_id="?",
                    name="?",
                    action="skipped", sheet=sheet_name,
                    message="نام و کد ملی هر دو خالی هستند",
                )
            # TEMP-NNNN-firstname-lastname  (max 30 chars safe)
            slug = f"{fn[:4]}{ln[:4]}".replace(" ","").upper() or "XX"
            national_id = f"TEMP{row_num:04d}{slug}"
            _nid_auto_generated = True

        name = f"{cell(COL['first_name']) or ''} {cell(COL['last_name']) or ''}".strip()

        # ── 2. Parse core fields ───────────────────────────────────
        first_name  = (cell(COL["first_name"]) or "").strip()
        last_name   = (cell(COL["last_name"]) or "").strip()
        father_name = (cell(COL["father_name"]) or "").strip()

        if not first_name or not last_name:
            return RowResult(
                row_num=row_num, national_id=national_id, name=name,
                action="skipped", sheet=sheet_name,
                message="نام یا نام خانوادگی خالی است",
            )

        dob = jalali_to_gregorian(cell(COL["dob"]))

        phone        = normalise_phone(cell(COL["phone"]))
        father_phone = normalise_phone(cell(COL["father_phone"]))
        mother_phone = normalise_phone(cell(COL["mother_phone"]))

        height = safe_int(cell(COL["height"]))
        weight = safe_decimal(cell(COL["weight"]))

        hand = map_hand_foot(cell(COL["hand"]))
        foot = map_hand_foot(cell(COL["foot"]))

        father_edu = map_education(cell(COL["father_edu"]))
        mother_edu = map_education(cell(COL["mother_edu"]))
        father_job = (cell(COL["father_job"]) or "").strip()
        mother_job = (cell(COL["mother_job"]) or "").strip()

        skill_level = (cell(COL["skill_level"]) or "").strip().upper()

        # ── 3. Insurance ───────────────────────────────────────────
        insurance_raw  = cell(COL["insurance"])
        ins_info       = detect_insurance(insurance_raw, insurance_fill)
        insurance_status = {
            "active":      "active",
            "expired":     "expired",
            "near_expiry": "active",   # still active, just soon-to-expire
            "none":        "none",
        }.get(ins_info.status, "none")

        # ── 4. Category (auto-create) ──────────────────────────────
        category_name = (cell(COL["category"]) or "").strip()
        category_obj  = None
        if category_name and not dry_run:
            category_obj, cat_created = self._get_or_create_category(
                name=category_name, result=result
            )
        elif category_name and dry_run:
            category_obj = None   # dry run — don't touch DB

        # ── 5. Dry run early exit ──────────────────────────────────
        if dry_run:
            return RowResult(
                row_num=row_num, national_id=national_id, name=name,
                action="skipped", sheet=sheet_name,
                message=f"[DRY RUN] دسته: {category_name} | بیمه: {ins_info.status}",
            )

        # ── 6. Upsert Player ──────────────────────────────────────
        try:
            from futsal_club.models import Player  # Django model import
            defaults = {
                "first_name":           first_name,
                "last_name":            last_name,
                "father_name":          father_name,
                "phone":                phone,
                "father_phone":         father_phone,
                "mother_phone":         mother_phone,
                "height":               height,
                "weight":               weight,
                "preferred_hand":       hand,
                "preferred_foot":       foot,
                "father_education":     father_edu,
                "mother_education":     mother_edu,
                "father_job":           father_job,
                "mother_job":           mother_job,
                "insurance_status":     insurance_status,
                "status":               Player.Status.APPROVED,
            }
            if dob:
                defaults["dob"] = dob
            if ins_info.expiry_date:
                defaults["insurance_expiry_date"] = ins_info.expiry_date

            player, created = Player.objects.update_or_create(
                national_id=national_id,
                defaults=defaults,
            )

            # Assign category M2M
            if category_obj:
                player.categories.add(category_obj)

            # Create/update TechnicalProfile for skill_level
            if skill_level:
                from futsal_club.models import TechnicalProfile
                TechnicalProfile.objects.update_or_create(
                    player=player,
                    defaults={"skill_level": skill_level, "updated_by": created_by},
                )

            action = "created" if created else "updated"
            nid_note = " [شناسه موقت]" if _nid_auto_generated else ""
            return RowResult(
                row_num=row_num, national_id=national_id, name=name,
                action=action, sheet=sheet_name,
                message=f"دسته: {category_name} | بیمه: {ins_info.status}{nid_note}",
            )

        except Exception as exc:
            logger.error("Row %d (%s): %s", row_num, national_id, exc, exc_info=True)
            return RowResult(
                row_num=row_num, national_id=national_id, name=name,
                action="error", sheet=sheet_name,
                message=str(exc)[:200],
            )

    # ── Category helper ────────────────────────────────────────────
    def _get_or_create_category(
        self, name: str, result: ImportResult
    ) -> Tuple:
        from futsal_club.models import TrainingCategory
        obj, created = TrainingCategory.objects.get_or_create(
            name=name,
            defaults={
                "is_active":    True,
                "monthly_fee":  0,
            }
        )
        if created:
            result.categories_created += 1
            result.warnings.append(f"دسته جدید ایجاد شد: «{name}»")
            logger.info("Auto-created TrainingCategory: %s", name)
        return obj, created


# ══════════════════════════════════════════════════════════════════════
#  STANDALONE RUNNER (for management command / testing without Django)
# ══════════════════════════════════════════════════════════════════════

def run_import(
    filepath: str,
    dry_run: bool = False,
    sheet_names: Optional[List[str]] = None,
) -> ImportResult:
    """
    Entry point for management command and Celery task.
    Django must be set up before calling this.
    """
    svc = ExcelImportService(filepath=filepath, sheet_names=sheet_names)
    return svc.run(dry_run=dry_run)